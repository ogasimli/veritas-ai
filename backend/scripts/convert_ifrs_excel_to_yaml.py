#!/usr/bin/env python3
"""Convert IFRS disclosure checklist from Excel to YAML format."""

import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml

# Excel file path
EXCEL_PATH = (
    Path.home() / "Downloads" / "IFRS_e_Check_2024_global_versionv1_5_FINAL 1.xlsm"
)
# Output YAML path
YAML_PATH = (
    Path(__file__).parent.parent
    / "agents"
    / "veritas_ai_agent"
    / "data"
    / "ifrs_disclosure_checklist.yaml"
)

# Standard names mapping
STANDARD_NAMES = {
    "IAS 1": "Presentation of Financial Statements",
    "IAS 2": "Inventories",
    "IAS 7": "Statement of Cash Flows",
    "IAS 8": "Accounting Policies, Changes in Accounting Estimates and Errors",
    "IAS 10": "Events after the Reporting Period",
    "IAS 12": "Income Taxes",
    "IAS 16": "Property, Plant and Equipment",
    "IAS 19": "Employee Benefits",
    "IAS 20": "Accounting for Government Grants and Disclosure of Government Assistance",
    "IAS 21": "The Effects of Changes in Foreign Exchange Rates",
    "IAS 23": "Borrowing Costs",
    "IAS 24": "Related Party Disclosures",
    "IAS 26": "Accounting and Reporting by Retirement Benefit Plans",
    "IAS 27": "Separate Financial Statements",
    "IAS 28": "Investments in Associates and Joint Ventures",
    "IAS 29": "Financial Reporting in Hyperinflationary Economies",
    "IAS 32": "Financial Instruments: Presentation",
    "IAS 33": "Earnings per Share",
    "IAS 34": "Interim Financial Reporting",
    "IAS 36": "Impairment of Assets",
    "IAS 37": "Provisions, Contingent Liabilities and Contingent Assets",
    "IAS 38": "Intangible Assets",
    "IAS 40": "Investment Property",
    "IAS 41": "Agriculture",
    "IFRS 1": "First-time Adoption of International Financial Reporting Standards",
    "IFRS 2": "Share-based Payment",
    "IFRS 3": "Business Combinations",
    "IFRS 4": "Insurance Contracts",
    "IFRS 5": "Non-current Assets Held for Sale and Discontinued Operations",
    "IFRS 6": "Exploration for and Evaluation of Mineral Resources",
    "IFRS 7": "Financial Instruments: Disclosures",
    "IFRS 8": "Operating Segments",
    "IFRS 9": "Financial Instruments",
    "IFRS 10": "Consolidated Financial Statements",
    "IFRS 11": "Joint Arrangements",
    "IFRS 12": "Disclosure of Interests in Other Entities",
    "IFRS 13": "Fair Value Measurement",
    "IFRS 14": "Regulatory Deferral Accounts",
    "IFRS 15": "Revenue from Contracts with Customers",
    "IFRS 16": "Leases",
    "IFRS 17": "Insurance Contracts",
    "IFRS 18": "Presentation and Disclosure in Financial Statements",
}


def extract_standard_from_reference(reference):
    """Extract standard code (e.g., 'IAS 1', 'IFRS 15') from reference string."""
    if not reference or not isinstance(reference, str):
        return None

    reference = reference.strip()

    # Direct standard mention (e.g., "IAS 1", "IFRS 15")
    match = re.match(r"^(IAS|IFRS)\s+(\d+)", reference, re.IGNORECASE)
    if match:
        return f"{match.group(1).upper()} {match.group(2)}"

    # Paragraph reference (e.g., "1p15" = IAS 1 paragraph 15, "15p10" = IFRS 15 paragraph 10)
    match = re.match(r"^(\d+)p\d+", reference)
    if match:
        std_num = match.group(1)
        # Determine if IAS or IFRS based on number
        # IAS: 1-41, IFRS: 1-18 (with some gaps)
        # This is a heuristic - both can have overlapping numbers
        # For numbers 1-10, could be either, so we need context
        # Let's assume IFRS for common ones: 9, 15, 16, 17
        if std_num in ["9", "15", "16", "17", "18"]:
            return f"IFRS {std_num}"
        else:
            return f"IAS {std_num}"

    # IFRIC or SIC references
    if reference.startswith(("IFRIC", "SIC")):
        # We'll skip these for now as they're interpretations
        return None

    return None


def clean_disclosure_text(text):
    """Clean and validate disclosure text."""
    if not text or not isinstance(text, str):
        return None

    text = text.strip()

    # Skip if too short
    if len(text) < 10:
        return None

    # Skip summary rows
    if "Total:" in text or "| P -" in text:
        return None

    # Skip if it's just a reference
    if re.match(r"^[\d\w]+p[\d\w,\s\-()]+$", text):
        return None

    return text


def convert_excel_to_yaml():
    """Convert Excel disclosure checklist to YAML format."""
    print(f"Loading Excel file: {EXCEL_PATH}")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")

    # Use the e-Check sheet
    checklist_sheet = wb["e-Check"]
    print("\nProcessing sheet: e-Check")

    # Data structure for output
    standards = {}
    disclosure_counter = defaultdict(int)

    # Track current standard for section headers
    current_standard = None

    # Process rows starting from row 10 (after headers)
    for _row_num, row in enumerate(
        checklist_sheet.iter_rows(min_row=10, values_only=True), 10
    ):
        if not any(row):  # Skip empty rows
            continue

        # Column C (index 2) = Reference, Column D (index 3) = Disclosure
        reference = row[2]
        disclosure = row[3]

        # Try to extract standard from reference
        std_code = extract_standard_from_reference(reference)

        # If we found a new standard, update current_standard
        if std_code:
            current_standard = std_code
            # Initialize standard if not exists
            if std_code not in standards:
                standards[std_code] = {
                    "name": STANDARD_NAMES.get(std_code, f"{std_code} Standard"),
                    "disclosures": [],
                }

        # Clean disclosure text
        clean_disc = clean_disclosure_text(disclosure)

        # Add disclosure if valid and we have a current standard
        if clean_disc and current_standard:
            disclosure_counter[current_standard] += 1
            disclosure_id = f"{current_standard.replace(' ', '')}-D{disclosure_counter[current_standard]}"

            # Create a shortened requirement title (first 100 chars or first sentence)
            requirement_title = (
                clean_disc[:100] if len(clean_disc) <= 100 else clean_disc[:97] + "..."
            )

            standards[current_standard]["disclosures"].append(
                {
                    "id": disclosure_id,
                    "requirement": requirement_title,
                    "description": clean_disc,
                }
            )

    # Filter out standards with very few disclosures (likely parsing errors)
    standards = {k: v for k, v in standards.items() if len(v["disclosures"]) >= 5}

    # Create output structure
    output = {"standards": standards}

    # Print summary
    print(f"\n{'=' * 60}")
    print("Conversion Summary:")
    print(f"{'=' * 60}")
    print(f"Total standards found: {len(standards)}")
    for std_code, std_data in sorted(standards.items(), key=lambda x: x[0]):
        print(
            f"  {std_code:10s} ({std_data['name'][:40]:40s}): {len(std_data['disclosures']):3d} disclosures"
        )

    # Save to YAML
    YAML_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(YAML_PATH, "w") as f:
        yaml.dump(
            output,
            f,
            default_flow_style=False,
            sort_keys=False,
            allow_unicode=True,
            width=120,
        )

    print(f"\nYAML file saved to: {YAML_PATH}")
    print("\nSample disclosure from IAS 1:")
    if "IAS 1" in standards and standards["IAS 1"]["disclosures"]:
        sample = standards["IAS 1"]["disclosures"][0]
        print(f"  ID: {sample['id']}")
        print(f"  Requirement: {sample['requirement']}")
        print(f"  Description: {sample['description'][:200]}...")


if __name__ == "__main__":
    convert_excel_to_yaml()
